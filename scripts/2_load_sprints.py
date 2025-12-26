# scripts/2_load_sprints_smart.py - Faqat yangi fayllarni yuklash
from openpyxl import load_workbook
import sys
import os
import numpy as np
from tqdm import tqdm
import json
from datetime import datetime

from utils.chunking_helper import ChunkingHelper

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.embedding_helper import EmbeddingHelper
from utils.vectordb_helper import VectorDBHelper
from dotenv import load_dotenv

load_dotenv()

print("=" * 80)
print("📊 EXCEL REPORTLARNI VECTORDB GA YUKLASH")
print("🎯 SMART CHUNKING + FAQAT YANGI FAYLLAR")
print("=" * 80)
print()

# Yuklangan fayllarni saqlash uchun log fayl
LOADED_FILES_LOG = "loaded_files.json"


def load_processed_files():
    """Allaqachon yuklangan fayllar ro'yxatini o'qish"""
    if os.path.exists(LOADED_FILES_LOG):
        try:
            with open(LOADED_FILES_LOG, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_processed_file(filename, file_info):
    """Yuklangan faylni log'ga qo'shish"""
    processed = load_processed_files()
    processed[filename] = file_info
    with open(LOADED_FILES_LOG, 'w', encoding='utf-8') as f:
        json.dump(processed, f, indent=2, ensure_ascii=False)


def get_file_hash(filepath):
    """Fayl o'zgarganligini tekshirish uchun hash (file size + modified time)"""
    stat = os.stat(filepath)
    return f"{stat.st_size}_{int(stat.st_mtime)}"


# 1. Helpers
print("📦 Helpers yuklanmoqda...")
embedding_helper = EmbeddingHelper()
vectordb_helper = VectorDBHelper()
chunking_helper = ChunkingHelper(max_chunk_length=1500)
print("✅ Tayyor!")
print()

# 2. Excel papkasi
excel_dir = os.getenv('EXCEL_DIR')
if not excel_dir or not os.path.exists(excel_dir):
    print(f"❌ Excel papkasi topilmadi: {excel_dir}")
    print("   .env faylingizda EXCEL_DIR ni to'g'ri ko'rsating")
    sys.exit(1)

excel_files = [
    f for f in os.listdir(excel_dir)
    if f.endswith('.xlsx') and not f.startswith('~$')
]

if not excel_files:
    print(f"⚠️  Excel fayllar topilmadi: {excel_dir}")
    sys.exit(1)

print(f"📁 Topildi: {len(excel_files)} ta Excel fayl")
print()

# 3. Allaqachon yuklangan fayllarni tekshirish
processed_files = load_processed_files()
new_files = []
skipped_files = []

print("🔍 Yangi fayllar tekshirilmoqda...")
for excel_file in excel_files:
    file_path = os.path.join(excel_dir, excel_file)
    file_hash = get_file_hash(file_path)

    if excel_file in processed_files:
        # Fayl allaqachon yuklangan, lekin o'zgarganmi?
        if processed_files[excel_file].get('hash') == file_hash:
            skipped_files.append(excel_file)
            print(f"   ⏭️  O'tkazib yuborildi: {excel_file} (allaqachon yuklangan)")
        else:
            new_files.append((excel_file, file_hash))
            print(f"   🔄 Yangilangan: {excel_file} (qayta yuklanadi)")
    else:
        new_files.append((excel_file, file_hash))
        print(f"   ✨ Yangi: {excel_file}")

print()
print(f"📊 Natija:")
print(f"   • Yangi/Yangilangan: {len(new_files)} ta")
print(f"   • O'tkazib yuborildi: {len(skipped_files)} ta")
print()

if not new_files:
    print("✅ Barcha fayllar allaqachon yuklangan!")
    print("=" * 80)
    sys.exit(0)

# 4. Faqat yangi fayllarni yuklash
total_loaded = 0
total_chunks = 0
total_root_causes = 0
total_solutions = 0

for file_idx, (excel_file, file_hash) in enumerate(new_files, 1):
    file_path = os.path.join(excel_dir, excel_file)

    print("=" * 80)
    print(f"📖 [{file_idx}/{len(new_files)}] {excel_file}")
    print("=" * 80)

    # Excel o'qish
    try:
        wb = load_workbook(file_path, read_only=False, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"❌ Faylni o'qishda xatolik: {e}")
        print()
        continue

    # Header nomlarini olish (birinchi qator)
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers[header] = col

    # Total rows count
    total_rows = ws.max_row - 1  # Minus header

    print(f"📋 Ustunlar: {len(headers)} ta")
    print(f"📊 Issues: {total_rows} ta")
    print(f"   Asosiy ustunlar: {', '.join(list(headers.keys())[:8])}...")
    print()

    # Ma'lumotlarni yig'ish
    keys = []
    weighted_embeddings = []
    full_texts = []
    metadatas = []
    all_chunks_data = []

    # Ma'lumotlarni o'qish (2-qatordan boshlab) - WITH PROGRESS BAR
    print("⏳ Ma'lumotlar o'qilmoqda...")

    with tqdm(total=total_rows, desc="   📖 Reading", unit="issue",
              bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]") as pbar:

        for row in range(2, ws.max_row + 1):
            # Key (A ustuni)
            key = ws.cell(row=row, column=headers.get('Key', 1)).value
            if not key:
                pbar.update(1)
                continue

            # Barcha maydonlarni olish
            summary = ws.cell(row=row, column=headers.get('Summary', 1)).value or ''
            description = ws.cell(row=row, column=headers.get('Description', 1)).value or ''
            issue_type = ws.cell(row=row, column=headers.get('Type', 1)).value or ''
            status = ws.cell(row=row, column=headers.get('Status', 1)).value or ''
            assignee = ws.cell(row=row, column=headers.get('Assignee', 1)).value or 'Unassigned'
            reporter = ws.cell(row=row, column=headers.get('Reporter', 1)).value or 'Unknown'
            priority = ws.cell(row=row, column=headers.get('Priority', 1)).value or 'None'
            story_points = ws.cell(row=row, column=headers.get('Story Points', 1)).value or ''
            created_date = ws.cell(row=row, column=headers.get('Created Date', 1)).value or ''
            resolved_date = ws.cell(row=row, column=headers.get('Resolved Date', 1)).value or ''

            comments = ws.cell(row=row, column=headers.get('Comments', 1)).value or ''
            comment_authors = ws.cell(row=row, column=headers.get('Comment Authors', 1)).value or ''
            return_count = ws.cell(row=row, column=headers.get('Return Count', 1)).value or 0
            return_reasons = ws.cell(row=row, column=headers.get('Return Reasons', 1)).value or ''

            status_history = ws.cell(row=row, column=headers.get('Status History', 1)).value or ''
            testing_time = ws.cell(row=row, column=headers.get('Testing Time', 1)).value or ''

            labels = ws.cell(row=row, column=headers.get('Labels', 1)).value or ''
            components = ws.cell(row=row, column=headers.get('Components', 1)).value or ''
            linked_issues = ws.cell(row=row, column=headers.get('Linked Issues', 1)).value or ''

            pr_status = ws.cell(row=row, column=headers.get('PR Status', 1)).value or ''
            pr_count = ws.cell(row=row, column=headers.get('PR Count', 1)).value or 0
            pr_last_updated = ws.cell(row=row, column=headers.get('PR Last Updated', 1)).value or ''

            # Sprint nomini fayldan ajratib olish
            parts = excel_file.replace('.xlsx', '').split('_')
            if 'Sprint' in excel_file:
                sprint_id = next((p for p in parts if p.isdigit()), "Unknown")
            else:
                sprint_id = "Unknown"

            # Issue data dictionary
            issue_data = {
                'key': key,
                'summary': summary,
                'description': description,
                'type': issue_type,
                'status': status,
                'assignee': assignee,
                'reporter': reporter,
                'priority': priority,
                'story_points': story_points,
                'created_date': str(created_date),
                'resolved_date': str(resolved_date),
                'comments': comments,
                'comment_authors': comment_authors,
                'return_count': return_count,
                'return_reasons': return_reasons,
                'status_history': status_history,
                'testing_time': testing_time,
                'labels': labels,
                'components': components,
                'linked_issues': linked_issues,
                'pr_status': pr_status,
                'pr_count': pr_count,
                'pr_last_updated': pr_last_updated,
                'sprint_id': sprint_id
            }

            # SMART CHUNKING
            chunks = chunking_helper.create_chunks(issue_data)
            total_chunks += len(chunks)

            # Statistika
            for chunk in chunks:
                if chunk['type'] == 'root_cause':
                    total_root_causes += 1
                elif chunk['type'] == 'solution':
                    total_solutions += 1

            # Full text
            full_text = chunking_helper.create_full_text_for_backward_compatibility(issue_data)

            # Metadata
            metadata = {
                'type': issue_type,
                'status': status,
                'sprint_id': sprint_id,
                'assignee': assignee,
                'reporter': reporter,
                'priority': priority,
                'story_points': str(story_points),
                'created_date': str(created_date),
                'resolved_date': str(resolved_date),
                'has_comments': 'yes' if comments else 'no',
                'return_count': str(return_count),
                'labels': labels if labels else 'none',
                'components': components if components else 'none',
                'has_pr': 'yes' if pr_status else 'no',
                'pr_status': pr_status if pr_status else 'none',
            }

            keys.append(key)
            full_texts.append(full_text)
            metadatas.append(metadata)
            all_chunks_data.append(chunks)

            pbar.update(1)

    wb.close()

    if not keys:
        print(f"   ⚠️  Ma'lumot topilmadi, o'tkazib yuborildi")
        print()
        continue

    print(f"✅ {len(keys)} ta issue o'qildi")
    print(f"   📦 Chunks: {sum(len(c) for c in all_chunks_data)} ta")
    print()

    # EMBEDDING - WITH PROGRESS BAR
    print("🔄 Embedding qilinmoqda...")
    print(f"   Strategy: Batch encoding (optimized)")

    # Barcha chunks'ni yig'ish
    all_chunks_flat = []
    chunk_counts = []

    for issue_chunks in all_chunks_data:
        chunk_counts.append(len(issue_chunks))
        all_chunks_flat.extend(issue_chunks)

    # BATCH ENCODING
    if all_chunks_flat:
        print(f"   ⚡ Batch size: {len(all_chunks_flat)} chunks")
        all_embeddings_flat = embedding_helper.encode_chunks(all_chunks_flat, show_progress=True)
        print(f"   ✅ Encoding tugadi!")
    else:
        all_embeddings_flat = []

    # Weighted average - WITH PROGRESS BAR
    print("   🧮 Weighted average hisoblash...")
    all_weighted_embeddings = []
    embedding_idx = 0

    with tqdm(total=len(chunk_counts), desc="   ⚖️  Weighting", unit="issue",
              bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt}") as pbar:

        for idx, chunk_count in enumerate(chunk_counts):
            if chunk_count > 0:
                issue_chunks = all_chunks_data[idx]
                issue_embeddings = all_embeddings_flat[embedding_idx:embedding_idx + chunk_count]
                embedding_idx += chunk_count

                for i, chunk in enumerate(issue_chunks):
                    chunk['embedding'] = issue_embeddings[i]

                weights = [chunk.get('weight', 1.0) for chunk in issue_chunks]
                total_weight = sum(weights)

                if total_weight > 0 and issue_embeddings:
                    weighted_embeddings_list = [
                        np.array(emb) * (weight / total_weight)
                        for emb, weight in zip(issue_embeddings, weights)
                    ]
                    weighted_average = np.sum(weighted_embeddings_list, axis=0).tolist()
                    all_weighted_embeddings.append(weighted_average)
                else:
                    all_weighted_embeddings.append([0.0] * 1024)
            else:
                all_weighted_embeddings.append([0.0] * 1024)

            pbar.update(1)

    print(f"   ✅ Weighted embeddings tayyor: {len(all_weighted_embeddings)}")
    print()

    # VectorDB - WITH ANIMATION
    print("💾 VectorDB ga yuklanmoqda...")
    try:
        # Simple spinner animation
        with tqdm(total=len(keys), desc="   💾 Saving", unit="issue",
                  bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt}") as pbar:

            vectordb_helper.add_issues_batch_with_chunks(
                keys=keys,
                weighted_embeddings=all_weighted_embeddings,
                full_texts=full_texts,
                metadatas=metadatas,
                all_chunks_data=all_chunks_data
            )
            pbar.update(len(keys))

        total_loaded += len(keys)
        print(f"✅ Yuklandi: {len(keys)} ta issue")

        # Faylni log'ga qo'shish
        file_info = {
            'hash': file_hash,
            'loaded_at': datetime.now().isoformat(),
            'issues_count': len(keys),
            'chunks_count': sum(len(c) for c in all_chunks_data)
        }
        save_processed_file(excel_file, file_info)
        print(f"   📝 Log'ga yozildi")

    except Exception as e:
        print(f"❌ VectorDB ga yuklashda xatolik: {e}")

    print()

# YAKUNIY STATISTIKA
print()
print("=" * 80)
print("🎉 YAKUNIY NATIJA")
print("=" * 80)

stats = vectordb_helper.get_stats()
print(f"📊 VectorDB:")
print(f"   • Jami issues: {stats['total_issues']} ta")
print(f"   • Yangi yuklandi: {total_loaded} ta")
print()

if total_loaded > 0:
    print(f"📦 Chunking:")
    print(f"   • Jami chunks: {total_chunks} ta")
    print(f"   • O'rtacha per issue: {total_chunks / total_loaded:.1f}")
    print()

    print(f"🎯 Smart Detection:")
    print(f"   • Root causes detected: {total_root_causes} ta")
    print(f"   • Solutions detected: {total_solutions} ta")
    detection_rate = ((total_root_causes + total_solutions) / total_loaded) * 100
    print(f"   • Detection rate: {detection_rate:.1f}%")
    print()

print(f"📁 Processed files log: {LOADED_FILES_LOG}")
print()
print("✅ TAYYOR!")
print("=" * 80)