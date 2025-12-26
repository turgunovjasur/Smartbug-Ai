"""
JIRA SPRINT REPORT - PULL REQUEST VA TESTING RETURN BILAN
---------------------------------------------------------
Yangi ustunlar:
✅ PR Status, PR Count, PR Last Updated
✅ Testing Return Who, Testing Return When
"""

from jira import JIRA
from dotenv import load_dotenv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from collections import defaultdict
from functools import lru_cache
from tqdm import tqdm
import logging
import json
import re

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

load_dotenv()


# ============================================================================
# CONFIGURATION
# ============================================================================
class Config:
    # Sprint va Project
    SPRINT_IDS = [2842, 3014, 2775, 2379, 2352, 2351, 2148]
    PROJECT_KEY = 'DEV'
    BOARD_ID = 10

    # Custom fields
    STORY_POINTS_FIELD = 'customfield_10016'
    SPRINT_FIELD = 'customfield_10020'
    PR_FIELD = 'customfield_10000'  # ← YANGI!

    # Status configuration
    TESTING_STATUSES = ['TESTING', 'Ready to Test', 'NEED CLARIFICATION/RETURN TEST']
    DONE_STATUSES = ['CLOSED', 'Closed', 'Done', 'Resolved']
    RETURN_STATUS = 'NEED CLARIFICATION/RETURN TEST'  # ← YANGI!

    # Excel styling
    HEADER_COLOR = "4472C4"
    CLOSED_STATUS_COLOR = "C6EFCE"


# ============================================================================
# USTUNLAR - Optimizatsiya qilingan (foydasizlar o'chirildi)
# ============================================================================
ACTIVE_COLUMNS = [
    'Key',
    'Sprint',
    'Summary',
    'Description',
    'Type',
    'Status',
    'Priority',
    'Assignee',
    'Reporter',
    'Story Points',
    'Created Date',
    'Resolved Date',
    'Added to Sprint',

    # Pull Request
    'PR Status',
    'PR Count',
    'PR Last Updated',

    # Comments
    'Comment Count',
    'Comments',
    'Comment Authors',

    # Status & Time
    'Status History',
    'Time in Each Status',
    'Testing Time',

    # Testing Returns
    'Return Count',
    'Return Reasons',
    'Testing Return Who',
    'Testing Return When',

    # Other
    'Linked Issues',
]

COLUMN_WIDTHS = {
    'Key': 12,
    'Sprint': 25,
    'Summary': 40,
    'Description': 50,
    'Type': 12,
    'Status': 15,
    'Priority': 12,
    'Assignee': 20,
    'Reporter': 20,
    'Story Points': 12,
    'Created Date': 15,
    'Resolved Date': 15,
    'Added to Sprint': 15,
    'PR Status': 12,
    'PR Count': 10,
    'PR Last Updated': 18,
    'Comment Count': 12,
    'Comments': 60,
    'Comment Authors': 30,
    'Status History': 60,
    'Time in Each Status': 60,
    'Testing Time': 15,
    'Return Count': 12,
    'Return Reasons': 50,
    'Testing Return Who': 30,
    'Testing Return When': 30,
    'Linked Issues': 30,
}


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================
def safe_get(obj, attr, default=''):
    """Xavfsiz attribute olish"""
    try:
        value = getattr(obj, attr, default)
        return value if value is not None else default
    except:
        return default


def safe_date(date_str):
    """Sanani formatlash"""
    try:
        return date_str[:10] if date_str else ''
    except:
        return ''


# ============================================================================
# PULL REQUEST EXTRACTION
# ============================================================================
def extract_pr_info(issue):
    """Pull Request ma'lumotlarini extract qilish"""
    pr_status = ''
    pr_count = 0
    pr_last_updated = ''

    try:
        pr_data = getattr(issue.fields, Config.PR_FIELD, None)

        if pr_data:
            # String bo'lsa, JSON parse qilish
            if isinstance(pr_data, str):
                try:
                    pr_json = json.loads(pr_data)
                except:
                    # JSON emas, odd string
                    return '', 0, ''
            elif isinstance(pr_data, dict):
                pr_json = pr_data
            else:
                return '', 0, ''

            # PR state ni olish
            if 'pullrequest' in pr_json:
                pr_info = pr_json['pullrequest']
                pr_status = pr_info.get('state', '')
                pr_count = pr_info.get('stateCount', 0)

            # Last updated ni olish
            if 'json' in pr_json:
                json_data = pr_json['json']
                if isinstance(json_data, str):
                    try:
                        json_data = json.loads(json_data)
                    except:
                        pass

                if isinstance(json_data, dict):
                    cached = json_data.get('cachedValue', {})
                    summary = cached.get('summary', {})
                    pr_overall = summary.get('pullrequest', {}).get('overall', {})
                    last_updated = pr_overall.get('lastUpdated', '')

                    if last_updated:
                        pr_last_updated = last_updated[:16].replace('T', ' ')
    except Exception as e:
        logger.debug(f"PR extraction error for {issue.key}: {e}")

    return pr_status, pr_count, pr_last_updated


# ============================================================================
# TESTING RETURN EXTRACTION
# ============================================================================
def extract_testing_returns_detailed(issue):
    """Testing dan qaytganlar - kim va qachon"""
    who_list = []
    when_list = []

    try:
        if not hasattr(issue, 'changelog') or not issue.changelog:
            return '', ''

        for history in issue.changelog.histories:
            for item in history.items:
                if item.field == 'status':
                    if item.fromString in Config.TESTING_STATUSES and item.toString == Config.RETURN_STATUS:
                        date = history.created[:16].replace('T', ' ')
                        who = safe_get(history.author, 'displayName', 'Unknown') if hasattr(history,
                                                                                            'author') else 'Unknown'

                        when_list.append(date)
                        who_list.append(who)
    except:
        pass

    return '\n'.join(who_list), '\n'.join(when_list)


# ============================================================================
# COLUMN EXTRACTORS
# ============================================================================
def extract_key(issue, sprint_info_map):
    return issue.key


def extract_sprint(issue, sprint_info_map):
    """Sprint va sprint qo'shilgan sana"""
    found_sprints = []
    sprint_added_date = ''

    try:
        if hasattr(issue.fields, Config.SPRINT_FIELD):
            sprint_field = getattr(issue.fields, Config.SPRINT_FIELD)
            if sprint_field and isinstance(sprint_field, list):
                for sprint_obj in sprint_field:
                    if hasattr(sprint_obj, 'id'):
                        obj_id = getattr(sprint_obj, 'id')
                        if obj_id in Config.SPRINT_IDS:
                            sprint_name = sprint_info_map.get(obj_id, {}).get('name', f'Sprint {obj_id}')
                            if sprint_name not in found_sprints:
                                found_sprints.append(sprint_name)
    except:
        pass

    try:
        if hasattr(issue, 'changelog') and issue.changelog:
            for history in issue.changelog.histories:
                for item in history.items:
                    if item.field == 'Sprint' and not sprint_added_date:
                        sprint_added_date = history.created[:10]
                        break
    except:
        pass

    if not found_sprints:
        found_sprints = ['Unknown']

    if not sprint_added_date:
        sprint_added_date = safe_date(issue.fields.created)

    return ', '.join(found_sprints), sprint_added_date


def extract_summary(issue, sprint_info_map):
    return safe_get(issue.fields, 'summary')


def extract_description(issue, sprint_info_map):
    desc = safe_get(issue.fields, 'description')
    return desc if desc else ''


def extract_type(issue, sprint_info_map):
    return safe_get(issue.fields.issuetype, 'name') if issue.fields.issuetype else ''


def extract_status(issue, sprint_info_map):
    return safe_get(issue.fields.status, 'name') if issue.fields.status else ''


def extract_priority(issue, sprint_info_map):
    return safe_get(issue.fields.priority, 'name', 'None') if issue.fields.priority else 'None'


def extract_assignee(issue, sprint_info_map):
    return safe_get(issue.fields.assignee, 'displayName', 'Unassigned') if issue.fields.assignee else 'Unassigned'


def extract_reporter(issue, sprint_info_map):
    return safe_get(issue.fields.reporter, 'displayName', 'Unknown') if issue.fields.reporter else 'Unknown'


def extract_story_points(issue, sprint_info_map):
    return safe_get(issue.fields, Config.STORY_POINTS_FIELD, '')


def extract_created_date(issue, sprint_info_map):
    return safe_date(issue.fields.created)


def extract_resolved_date(issue, sprint_info_map):
    return safe_date(issue.fields.resolutiondate)


def extract_pr_status(issue, sprint_info_map):
    """PR Status"""
    status, _, _ = extract_pr_info(issue)
    return status


def extract_pr_count(issue, sprint_info_map):
    """PR Count"""
    _, count, _ = extract_pr_info(issue)
    return count


def extract_pr_last_updated(issue, sprint_info_map):
    """PR Last Updated"""
    _, _, last_updated = extract_pr_info(issue)
    return last_updated


def extract_comment_count(issue, sprint_info_map):
    try:
        if hasattr(issue.fields, 'comment') and hasattr(issue.fields.comment, 'comments'):
            return len(issue.fields.comment.comments)
        return 0
    except:
        return 0


def extract_comments(issue, sprint_info_map):
    try:
        if hasattr(issue.fields, 'comment') and hasattr(issue.fields.comment, 'comments'):
            comments = issue.fields.comment.comments
            texts = []
            for c in comments:
                if hasattr(c, 'created') and hasattr(c, 'body') and c.body:
                    date = c.created[:10]
                    texts.append(f"[{date}] {c.body}")
            return '\n\n'.join(texts)
        return ''
    except:
        return ''


def extract_comment_authors(issue, sprint_info_map):
    try:
        if hasattr(issue.fields, 'comment') and hasattr(issue.fields.comment, 'comments'):
            comments = issue.fields.comment.comments
            authors = []
            for c in comments:
                if hasattr(c, 'author') and hasattr(c.author, 'displayName'):
                    author = c.author.displayName
                    if author not in authors:
                        authors.append(author)
            return ', '.join(authors)
        return ''
    except:
        return ''


def extract_status_history(issue, sprint_info_map):
    history = []
    try:
        if hasattr(issue, 'changelog') and issue.changelog:
            for h in issue.changelog.histories:
                for item in h.items:
                    if item.field == 'status':
                        date = h.created[:16].replace('T', ' ')
                        from_status = item.fromString or 'None'
                        to_status = item.toString or 'None'
                        history.append(f"{date}: {from_status} → {to_status}")
    except:
        pass
    return '\n'.join(history)


def extract_time_in_each_status(issue, sprint_info_map):
    status_durations = {}

    try:
        if not hasattr(issue, 'changelog') or not issue.changelog:
            return ''

        status_changes = []
        for history in issue.changelog.histories:
            for item in history.items:
                if item.field == 'status':
                    status_changes.append({
                        'date': history.created,
                        'from': item.fromString or 'None',
                        'to': item.toString or 'None'
                    })

        if not status_changes:
            return ''

        # Boshlang'ich status
        initial_status = status_changes[0]['from']
        initial_date = datetime.fromisoformat(status_changes[0]['date'].replace('Z', '+00:00'))
        created_date = datetime.fromisoformat(issue.fields.created.replace('Z', '+00:00'))
        duration = (initial_date - created_date).total_seconds() / 3600
        status_durations[initial_status] = duration

        # Har bir status
        for i in range(len(status_changes)):
            current_status = status_changes[i]['to']
            current_date = datetime.fromisoformat(status_changes[i]['date'].replace('Z', '+00:00'))

            if i + 1 < len(status_changes):
                next_date = datetime.fromisoformat(status_changes[i + 1]['date'].replace('Z', '+00:00'))
            else:
                if issue.fields.resolutiondate:
                    next_date = datetime.fromisoformat(issue.fields.resolutiondate.replace('Z', '+00:00'))
                else:
                    next_date = datetime.now(current_date.tzinfo)

            duration = (next_date - current_date).total_seconds() / 3600

            if current_status in status_durations:
                status_durations[current_status] += duration
            else:
                status_durations[current_status] = duration

        # Format
        result = []
        for status, hours in sorted(status_durations.items(), key=lambda x: x[1], reverse=True):
            result.append(f"{status}: {hours:.1f}h ({hours / 24:.1f}d)")

        return '\n'.join(result)
    except:
        return ''


def extract_testing_time(issue, sprint_info_map):
    total_hours = 0

    try:
        if not hasattr(issue, 'changelog') or not issue.changelog:
            return ''

        status_changes = []
        for history in issue.changelog.histories:
            for item in history.items:
                if item.field == 'status':
                    status_changes.append({
                        'date': history.created,
                        'to': item.toString or 'None'
                    })

        for i, change in enumerate(status_changes):
            if change['to'] in Config.TESTING_STATUSES:
                start = datetime.fromisoformat(change['date'].replace('Z', '+00:00'))

                if i + 1 < len(status_changes):
                    end = datetime.fromisoformat(status_changes[i + 1]['date'].replace('Z', '+00:00'))
                else:
                    if issue.fields.resolutiondate:
                        end = datetime.fromisoformat(issue.fields.resolutiondate.replace('Z', '+00:00'))
                    else:
                        end = datetime.now(start.tzinfo)

                hours = (end - start).total_seconds() / 3600
                total_hours += hours
    except:
        pass

    return f"{total_hours:.1f}h" if total_hours > 0 else ''


def extract_return_count(issue, sprint_info_map):
    return_count = 0

    try:
        if not hasattr(issue, 'changelog') or not issue.changelog:
            return 0

        for history in issue.changelog.histories:
            for item in history.items:
                if item.field == 'status':
                    if item.fromString in Config.TESTING_STATUSES and item.toString == Config.RETURN_STATUS:
                        return_count += 1
    except:
        pass

    return return_count


def extract_return_reasons(issue, sprint_info_map):
    return_reasons = []
    return_count = 0

    try:
        if not hasattr(issue, 'changelog') or not issue.changelog:
            return ''

        status_changes = []
        for history in issue.changelog.histories:
            for item in history.items:
                if item.field == 'status':
                    status_changes.append({
                        'date': history.created[:16].replace('T', ' '),
                        'from': item.fromString or 'None',
                        'to': item.toString or 'None',
                        'author': safe_get(history.author, 'displayName', 'Unknown') if hasattr(history,
                                                                                                'author') else 'Unknown'
                    })

        for change in status_changes:
            if change['from'] in Config.TESTING_STATUSES and change['to'] == Config.RETURN_STATUS:
                return_count += 1
                return_reasons.append(
                    f"Return #{return_count} [{change['date']}]: {change['from']} → {change['to']} (by {change['author']})"
                )
    except:
        pass

    return '\n'.join(return_reasons)


def extract_testing_return_who(issue, sprint_info_map):
    """Kim qaytardi"""
    who, _ = extract_testing_returns_detailed(issue)
    return who


def extract_testing_return_when(issue, sprint_info_map):
    """Qachon qaytdi"""
    _, when = extract_testing_returns_detailed(issue)
    return when


def extract_linked_issues(issue, sprint_info_map):
    linked = []
    try:
        if issue.fields.issuelinks:
            for link in issue.fields.issuelinks:
                if hasattr(link, 'inwardIssue'):
                    linked.append(link.inwardIssue.key)
                elif hasattr(link, 'outwardIssue'):
                    linked.append(link.outwardIssue.key)
    except:
        pass
    return ', '.join(linked)


# ============================================================================
# COLUMN MAPPING
# ============================================================================
COLUMN_FUNCTIONS = {
    'Key': extract_key,
    'Sprint': lambda issue, si: extract_sprint(issue, si)[0],
    'Summary': extract_summary,
    'Description': extract_description,
    'Type': extract_type,
    'Status': extract_status,
    'Priority': extract_priority,
    'Assignee': extract_assignee,
    'Reporter': extract_reporter,
    'Story Points': extract_story_points,
    'Created Date': extract_created_date,
    'Resolved Date': extract_resolved_date,
    'Added to Sprint': lambda issue, si: extract_sprint(issue, si)[1],
    'PR Status': extract_pr_status,
    'PR Count': extract_pr_count,
    'PR Last Updated': extract_pr_last_updated,
    'Comment Count': extract_comment_count,
    'Comments': extract_comments,
    'Comment Authors': extract_comment_authors,
    'Status History': extract_status_history,
    'Time in Each Status': extract_time_in_each_status,
    'Testing Time': extract_testing_time,
    'Return Count': extract_return_count,
    'Return Reasons': extract_return_reasons,
    'Testing Return Who': extract_testing_return_who,
    'Testing Return When': extract_testing_return_when,
    'Linked Issues': extract_linked_issues,
}


# ============================================================================
# JIRA CONNECTION
# ============================================================================
def get_jira_client():
    try:
        jira = JIRA(
            server=os.getenv('JIRA_SERVER'),
            basic_auth=(os.getenv('JIRA_EMAIL'), os.getenv('JIRA_API_TOKEN'))
        )
        logger.info(f"✅ Jira ga ulandi: {os.getenv('JIRA_SERVER')}")
        return jira
    except Exception as e:
        logger.error(f"❌ Jira ga ulanishda xatolik: {e}")
        raise


@lru_cache(maxsize=100)
def get_sprint_info(jira, sprint_id):
    try:
        sprint = jira.sprint(sprint_id)
        return {
            'id': sprint_id,
            'name': sprint.name,
            'state': sprint.state,
            'startDate': sprint.startDate[:10] if hasattr(sprint, 'startDate') and sprint.startDate else None,
            'endDate': sprint.endDate[:10] if hasattr(sprint, 'endDate') and sprint.endDate else None,
        }
    except Exception as e:
        logger.warning(f"Sprint {sprint_id} ma'lumotlarini olishda xatolik: {e}")
        return {
            'id': sprint_id,
            'name': f"Sprint {sprint_id}",
            'state': 'Unknown',
            'startDate': None,
            'endDate': None,
        }


# ============================================================================
# DATA FETCHING
# ============================================================================
def fetch_issues(jira, jql):
    logger.info("📥 Issuelarni yuklamoqda...")

    try:
        temp = jira.search_issues(jql, maxResults=0)
        total = temp.total if hasattr(temp, 'total') else 0
    except:
        temp = jira.search_issues(jql, maxResults=1)
        total = temp.total if hasattr(temp, 'total') else 1

    logger.info(f"Jami {total} ta issue topildi")

    print("   ⏳ Barcha issuelar yuklanmoqda...")
    all_issues = jira.search_issues(
        jql,
        maxResults=False,
        expand='changelog,renderedFields'
    )

    logger.info(f"✅ {len(all_issues)} ta issue yuklandi")
    return all_issues


# ============================================================================
# EXCEL GENERATION
# ============================================================================
def create_excel_report(issues, sprint_info_map, project_key):
    logger.info("📄 Excel yaratilmoqda...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sprint Report"

    # Styling
    header_fill = PatternFill(start_color=Config.HEADER_COLOR, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Headers
    for col_idx, column_name in enumerate(ACTIVE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

        # Column width
        width = COLUMN_WIDTHS.get(column_name, 20)
        if col_idx <= 26:
            col_letter = chr(64 + col_idx)
        else:
            col_letter = f"A{chr(64 + col_idx - 26)}"
        ws.column_dimensions[col_letter].width = width

    # Data rows
    logger.info(f"Ma'lumotlar yozilmoqda: {len(issues)} ta issue")

    with tqdm(total=len(issues), desc="Excel ga yozish") as pbar:
        for row_idx, issue in enumerate(issues, start=2):
            for col_idx, column_name in enumerate(ACTIVE_COLUMNS, 1):
                try:
                    func = COLUMN_FUNCTIONS.get(column_name)
                    if func:
                        value = func(issue, sprint_info_map)
                    else:
                        value = ''

                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

                    # Status rang
                    if column_name == 'Status' and value in Config.DONE_STATUSES:
                        cell.fill = PatternFill(start_color=Config.CLOSED_STATUS_COLOR, fill_type="solid")
                except Exception as e:
                    logger.debug(f"Ustun {column_name} uchun xatolik: {e}")
                    ws.cell(row=row_idx, column=col_idx, value='')

            pbar.update(1)

    # Auto filter va freeze
    last_col_idx = len(ACTIVE_COLUMNS)
    if last_col_idx <= 26:
        last_col = chr(64 + last_col_idx)
    else:
        last_col = f"A{chr(64 + last_col_idx - 26)}"
    ws.auto_filter.ref = f"A1:{last_col}{len(issues) + 1}"
    ws.freeze_panes = 'A2'

    return wb


# ============================================================================
# STATISTICS
# ============================================================================
def generate_statistics(issues, sprint_info_map):
    logger.info("📊 Statistika hisoblanmoqda...")

    stats = {
        'sprint': defaultdict(int),
        'developer': defaultdict(lambda: {'total': 0, 'closed': 0, 'bugs': 0, 'points': 0}),
        'type': defaultdict(int),
        'status': defaultdict(int),
        'bug_points': 0,
        'total_points': 0,
        'return_count': 0,
        'pr_count': 0,
        'testing_returner': defaultdict(int),
    }

    for issue in tqdm(issues, desc="Statistika"):
        sprint_name, _ = extract_sprint(issue, sprint_info_map)
        for sprint in sprint_name.split(', '):
            stats['sprint'][sprint] += 1

        assignee = extract_assignee(issue, sprint_info_map)
        stats['developer'][assignee]['total'] += 1

        status = extract_status(issue, sprint_info_map)
        if status in Config.DONE_STATUSES:
            stats['developer'][assignee]['closed'] += 1

        issue_type = extract_type(issue, sprint_info_map)
        if issue_type == 'Bug':
            stats['developer'][assignee]['bugs'] += 1

        sp = extract_story_points(issue, sprint_info_map)
        if sp:
            try:
                points = float(sp)
                stats['developer'][assignee]['points'] += points
                stats['total_points'] += points

                if issue_type == 'Bug':
                    stats['bug_points'] += points
            except:
                pass

        stats['type'][issue_type] += 1
        stats['status'][status] += 1

        return_count = extract_return_count(issue, sprint_info_map)
        stats['return_count'] += return_count

        # PR statistics
        pr_count = extract_pr_count(issue, sprint_info_map)
        if pr_count > 0:
            stats['pr_count'] += 1

        # Testing returner statistics
        who = extract_testing_return_who(issue, sprint_info_map)
        if who:
            for person in who.split('\n'):
                stats['testing_returner'][person] += 1

    return stats


def print_statistics(stats, total_issues):
    print("\n" + "=" * 80)
    print("📊 STATISTIKA")
    print("=" * 80)

    # Sprint bo'yicha
    print("\n📊 SPRINT BO'YICHA:")
    for sprint_name, count in sorted(stats['sprint'].items()):
        pct = (count / total_issues) * 100
        print(f"   {sprint_name:30s}: {count:4d} ta ({pct:5.1f}%)")

    # Status bo'yicha
    print("\n📈 STATUS BO'YICHA:")
    for status, count in sorted(stats['status'].items(), key=lambda x: x[1], reverse=True):
        pct = (count / total_issues) * 100
        bar = "█" * int(pct / 5)
        print(f"   {status:30s}: {count:4d} ta {bar} ({pct:5.1f}%)")

    # Developer bo'yicha
    print("\n👥 DEVELOPER BO'YICHA (TOP 10):")
    print(f"{'Developer':<25} {'Total':<8} {'Closed':<8} {'Bugs':<8} {'Points':<8}")
    print("-" * 65)

    top_devs = sorted(stats['developer'].items(), key=lambda x: x[1]['total'], reverse=True)[:10]
    for dev, dev_stats in top_devs:
        print(f"{dev:<25} {dev_stats['total']:<8} {dev_stats['closed']:<8} "
              f"{dev_stats['bugs']:<8} {dev_stats['points']:<8.1f}")

    # Type bo'yicha
    print("\n📋 TYPE BO'YICHA:")
    for issue_type, count in sorted(stats['type'].items(), key=lambda x: x[1], reverse=True):
        pct = (count / total_issues) * 100
        bar = "█" * int(pct / 5)
        print(f"   {issue_type:20s}: {count:3d} ta {bar} ({pct:.1f}%)")

    # Pull Request
    print("\n🔗 PULL REQUEST:")
    pr_pct = (stats['pr_count'] / total_issues) * 100 if total_issues > 0 else 0
    print(f"   PR bor tasklar: {stats['pr_count']}/{total_issues} ({pr_pct:.1f}%)")

    # Testing Returner
    print("\n🔄 KIM QAYTARDI (QA):")
    for returner, count in sorted(stats['testing_returner'].items(), key=lambda x: x[1], reverse=True):
        print(f"   {returner:30s}: {count} marta")

    # Bug statistikasi
    print("\n🐛 BUG STATISTIKASI:")
    if stats['total_points'] > 0:
        bug_pct = (stats['bug_points'] / stats['total_points']) * 100
        print(f"   Bug Points: {stats['bug_points']:.1f}")
        print(f"   Total Points: {stats['total_points']:.1f}")
        print(f"   Bug %: {bug_pct:.1f}%")

    # Return statistikasi
    print("\n🔄 TESTDAN QAYTGAN:")
    print(f"   Jami: {stats['return_count']} marta")
    if total_issues > 0:
        avg_return = stats['return_count'] / total_issues
        print(f"   O'rtacha: {avg_return:.2f} marta/task")


# ============================================================================
# MAIN
# ============================================================================

# ============================================================================
# MAIN - HAR BIR SPRINT UCHUN ALOHIDA
# ============================================================================
def main():
    print("=" * 80)
    print("🚀 JIRA REPORT - HAR BIR SPRINT UCHUN ALOHIDA EXCEL")
    print("=" * 80)
    print(f"📊 Aktiv ustunlar: {len(ACTIVE_COLUMNS)} ta")
    print(f"🏃 Sprintlar soni: {len(Config.SPRINT_IDS)} ta")
    print("=" * 80)
    print()

    jira = get_jira_client()

    # Sprint ma'lumotlarini olish
    sprint_info_map = {}
    for sprint_id in Config.SPRINT_IDS:
        sprint_info_map[sprint_id] = get_sprint_info(jira, sprint_id)

    print("\n📋 SPRINT MA'LUMOTLARI:")
    for sprint_id, info in sprint_info_map.items():
        print(f"   🏃 {info['name']} (ID: {sprint_id}, Status: {info['state']})")

    # Har bir sprint uchun alohida Excel yaratish
    for sprint_id in Config.SPRINT_IDS:
        print("\n" + "=" * 80)
        sprint_info = sprint_info_map[sprint_id]
        sprint_name = sprint_info['name']
        print(f"🔄 {sprint_name} ishlanmoqda...")
        print("=" * 80)

        # Faqat shu sprint uchun JQL
        jql = f'project = "{Config.PROJECT_KEY}" AND sprint = {sprint_id} ORDER BY created DESC'
        print(f"🔍 JQL: {jql}")

        # Issues olish
        issues = fetch_issues(jira, jql)

        if not issues:
            print(f"⚠️ {sprint_name} uchun issue topilmadi, o'tkazib yuborildi")
            continue

        # Statistika
        stats = generate_statistics(issues, {sprint_id: sprint_info})

        # Excel yaratish
        wb = create_excel_report(issues, {sprint_id: sprint_info}, Config.PROJECT_KEY)

        # Fayl nomi - sprint ID va nomi bilan
        safe_sprint_name = sprint_name.replace('/', '-').replace('\\', '-')
        filename = f'{Config.PROJECT_KEY}_Sprint_{sprint_id}_{safe_sprint_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)

        # Statistika chiqarish
        print_statistics(stats, len(issues))

        print("\n" + "=" * 80)
        print("✅ TAYYOR!")
        print("=" * 80)
        print(f"   📄 Fayl: {filename}")
        print(f"   📊 Issues: {len(issues)} ta")
        print(f"   📋 Ustunlar: {len(ACTIVE_COLUMNS)} ta")
        print("=" * 80)

    print("\n" + "=" * 80)
    print("🎉 BARCHA SPRINTLAR MUVAFFAQIYATLI YARATILDI!")
    print("=" * 80)
    print(f"   📊 Jami {len(Config.SPRINT_IDS)} ta Excel fayl yaratildi")
    print("=" * 80)


if __name__ == '__main__':
    main()