# ui/statistics.py
import streamlit as st
import plotly.express as px
import pandas as pd
from openpyxl import load_workbook
import os
from dotenv import load_dotenv
import time
import sys
from typing import Dict, List

from ui.styles import CHART_COLORS

load_dotenv()


def debug_log(message):
    """Real-time debug log"""
    print(f"[DEBUG] {message}", flush=True)
    sys.stdout.flush()


# ============================================================================
# SPRINT METADATA - FAQAT METADATA NI O'QISH (FAST!)
# ============================================================================

@st.cache_data(ttl=3600)
def get_sprint_metadata(excel_dir: str) -> Dict[str, dict]:
    """
    Faqat sprint metadata ni o'qish - JUDA TEZ!

    Returns:
        {
            'sprint_id': {
                'file': 'filename.xlsx',
                'total_rows': 150,
                'file_size': '2.5 MB'
            }
        }
    """
    debug_log("Sprint metadata olinmoqda...")

    if not os.path.exists(excel_dir):
        return {}

    excel_files = [
        f for f in os.listdir(excel_dir)
        if f.endswith('.xlsx') and not f.startswith('~$')
    ]

    if not excel_files:
        return {}

    metadata = {}

    for excel_file in excel_files:
        file_path = os.path.join(excel_dir, excel_file)

        try:
            # Sprint ID ni fayldan extract qilish
            parts = excel_file.replace('.xlsx', '').split('_')
            sprint_id = "Unknown"

            # Format: DEV_Report_PR_SPRINTID_DATE.xlsx
            for i, part in enumerate(parts):
                if part.isdigit() and len(part) == 4:  # Sprint ID
                    sprint_id = part
                    break

            # Faqat metadata - FAYL OCHMASDAN
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)

            # Row count - faqat tez hisoblash
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            total_rows = ws.max_row - 1  # Minus header
            wb.close()

            metadata[sprint_id] = {
                'file': excel_file,
                'total_rows': total_rows,
                'file_size': f"{file_size_mb:.1f} MB",
                'file_path': file_path
            }

            debug_log(f"  Sprint {sprint_id}: {total_rows} ta issue ({file_size_mb:.1f} MB)")

        except Exception as e:
            debug_log(f"  ERROR: {excel_file} - {str(e)}")
            continue

    debug_log(f"✅ {len(metadata)} ta sprint topildi")
    return metadata


# ============================================================================
# SINGLE SPRINT LOADER - FAQAT BITTA SPRINTNI YUKLASH
# ============================================================================

@st.cache_data(ttl=3600)
def load_single_sprint(file_path: str, sprint_id: str) -> pd.DataFrame:
    """
    Faqat bitta sprintni yuklash - OPTIMIZED!

    Args:
        file_path: Excel fayl yo'li
        sprint_id: Sprint ID

    Returns:
        DataFrame with all issues from this sprint
    """
    debug_log(f"Sprint {sprint_id} yuklanmoqda...")
    start_time = time.time()

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers[header] = col

        # Data extraction - OPTIMIZED
        data = []

        # Batch processing - har 50 ta row bir batch
        batch_size = 50
        total_rows = ws.max_row - 1

        for batch_start in range(2, ws.max_row + 1, batch_size):
            batch_end = min(batch_start + batch_size, ws.max_row + 1)

            for row_idx in range(batch_start, batch_end):
                try:
                    key = ws.cell(row_idx, headers.get('Key', 1)).value
                    if not key:
                        continue

                    # Helper function
                    def get_val(col_name, default=''):
                        col_idx = headers.get(col_name)
                        if col_idx:
                            val = ws.cell(row_idx, col_idx).value
                            return str(val) if val is not None else default
                        return default

                    # Return Count
                    return_count = 0
                    for col in ['Return Count', 'Return from Test Count']:
                        col_idx = headers.get(col)
                        if col_idx:
                            val = ws.cell(row_idx, col_idx).value
                            if val and str(val).isdigit():
                                return_count = int(val)
                                break

                    # Story Points
                    sp_col = headers.get('Story Points')
                    sp_val = ws.cell(row_idx, sp_col).value if sp_col else 0
                    story_points = float(sp_val) if sp_val and str(sp_val).replace('.', '').isdigit() else 0

                    data.append({
                        'key': str(key),
                        'sprint': sprint_id,
                        'summary': get_val('Summary', ''),
                        'type': get_val('Type', ''),
                        'status': get_val('Status', ''),
                        'assignee': get_val('Assignee', 'Unassigned'),
                        'reporter': get_val('Reporter', 'Unknown'),
                        'priority': get_val('Priority', 'None'),
                        'story_points': story_points,
                        'return_count': return_count,
                        'created_date': get_val('Created Date', '')[:10],
                        'resolved_date': get_val('Resolved Date', '')[:10],
                        'components': get_val('Components', ''),
                        'labels': get_val('Labels', ''),
                    })

                except:
                    continue

        wb.close()

        df = pd.DataFrame(data)

        # Data types
        df['story_points'] = pd.to_numeric(df['story_points'], errors='coerce').fillna(0)
        df['return_count'] = pd.to_numeric(df['return_count'], errors='coerce').fillna(0).astype(int)

        load_time = time.time() - start_time
        debug_log(f"✅ Sprint {sprint_id}: {len(df)} ta issue yuklandi ({load_time:.1f}s)")

        return df

    except Exception as e:
        debug_log(f"❌ ERROR: Sprint {sprint_id} - {str(e)}")
        return pd.DataFrame()


# ============================================================================
# MULTI-SPRINT LOADER - TANLANGAN SPRINTLARNI YUKLASH
# ============================================================================

def load_selected_sprints(metadata: Dict[str, dict], selected_sprints: List[str]) -> pd.DataFrame:
    """
    Tanlangan sprintlarni yuklash - LAZY LOADING

    Args:
        metadata: Sprint metadata
        selected_sprints: Tanlangan sprint ID'lar

    Returns:
        Combined DataFrame
    """
    if not selected_sprints:
        return pd.DataFrame()

    debug_log(f"Yuklash boshlandi: {len(selected_sprints)} ta sprint")

    # Progress UI
    progress_text = st.empty()
    progress_bar = st.progress(0)

    all_dfs = []

    for idx, sprint_id in enumerate(selected_sprints):
        progress_text.text(f"📥 [{idx + 1}/{len(selected_sprints)}] Sprint {sprint_id} yuklanmoqda...")
        progress_bar.progress((idx + 1) / len(selected_sprints))

        sprint_meta = metadata.get(sprint_id)
        if not sprint_meta:
            continue

        # Load sprint - CACHED!
        df = load_single_sprint(sprint_meta['file_path'], sprint_id)

        if not df.empty:
            all_dfs.append(df)

    # Clear progress
    progress_text.empty()
    progress_bar.empty()

    if not all_dfs:
        return pd.DataFrame()

    # Combine
    combined_df = pd.concat(all_dfs, ignore_index=True)

    debug_log(f"✅ Jami yuklandi: {len(combined_df)} ta issue")

    return combined_df


# ============================================================================
# VISUALIZATION FUNCTIONS - SIMPLIFIED
# ============================================================================

def render_overview_metrics(df):
    """Umumiy ko'rsatkichlar"""
    st.markdown("### 📊 Umumiy Ko'rsatkichlar")

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("Jami Issue", len(df))

    with col2:
        closed_statuses = ['Closed', 'CLOSED', 'Done', 'Resolved']
        closed_count = len(df[df['status'].isin(closed_statuses)])
        st.metric("Closed", closed_count, f"{closed_count / len(df) * 100:.0f}%" if len(df) > 0 else "0%")

    with col3:
        bug_count = len(df[df['type'] == 'Bug'])
        st.metric("Bugs", bug_count)

    with col4:
        total_sp = df['story_points'].sum()
        st.metric("Story Points", f"{total_sp:.0f}")

    with col5:
        total_returns = int(df['return_count'].sum())
        st.metric("Returns", total_returns)


def render_developers_tab(df):
    """Developers tab"""
    st.markdown("### 👥 Developer Performance")

    closed_statuses = ['Closed', 'CLOSED', 'Done', 'Resolved']

    dev_stats = df.groupby('assignee').agg({
        'key': 'count',
        'status': lambda x: sum(s in closed_statuses for s in x),
        'type': lambda x: (x == 'Bug').sum(),
        'return_count': 'sum',
        'story_points': 'sum'
    }).reset_index()

    dev_stats.columns = ['Developer', 'Total Tasks', 'Closed', 'Bugs', 'Returns', 'Story Points']
    dev_stats = dev_stats[dev_stats['Developer'] != 'Unassigned']
    dev_stats = dev_stats.sort_values('Total Tasks', ascending=False)

    col1, col2 = st.columns(2)

    with col1:
        fig = px.bar(
            dev_stats.head(10), x='Total Tasks', y='Developer', orientation='h',
            color='Total Tasks',
            color_continuous_scale=[[0, CHART_COLORS['primary']], [1, CHART_COLORS['secondary']]],
            title='📊 Tasklar (Top 10)'
        )
        fig.update_layout(
            paper_bgcolor=CHART_COLORS['bg'], plot_bgcolor=CHART_COLORS['bg'],
            font_color=CHART_COLORS['text'], showlegend=False,
            yaxis={'categoryorder': 'total ascending'}
        )
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        dev_stats['Completion Rate'] = (dev_stats['Closed'] / dev_stats['Total Tasks'] * 100).round(1)
        fig = px.bar(
            dev_stats[dev_stats['Total Tasks'] >= 3].head(10),
            x='Completion Rate', y='Developer', orientation='h',
            color='Completion Rate',
            color_continuous_scale=[[0, CHART_COLORS['danger']], [0.5, CHART_COLORS['warning']],
                                    [1, CHART_COLORS['success']]],
            title='✅ Completion Rate %'
        )
        fig.update_layout(
            paper_bgcolor=CHART_COLORS['bg'], plot_bgcolor=CHART_COLORS['bg'],
            font_color=CHART_COLORS['text'], showlegend=False,
            yaxis={'categoryorder': 'total ascending'}
        )
        st.plotly_chart(fig, use_container_width=True)

    st.dataframe(dev_stats, use_container_width=True, hide_index=True)


def render_bugs_tab(df):
    """Bugs tab - SIMPLIFIED"""
    st.markdown("### 🐛 Bug Analysis")
    bugs_df = df[df['type'] == 'Bug']

    if bugs_df.empty:
        st.success("✅ Bug topilmadi!")
        return

    col1, col2 = st.columns(2)

    with col1:
        bug_status = bugs_df['status'].value_counts().reset_index()
        bug_status.columns = ['Status', 'Count']
        fig = px.pie(bug_status, values='Count', names='Status', title='🎯 Bug Status', hole=0.4)
        fig.update_layout(paper_bgcolor=CHART_COLORS['bg'], font_color=CHART_COLORS['text'])
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        bug_priority = bugs_df['priority'].value_counts().reset_index()
        bug_priority.columns = ['Priority', 'Count']
        fig = px.bar(bug_priority, x='Priority', y='Count', title='⚡ Bug Priority')
        fig.update_layout(paper_bgcolor=CHART_COLORS['bg'], font_color=CHART_COLORS['text'])
        st.plotly_chart(fig, use_container_width=True)


def render_returns_tab(df):
    """Returns tab"""
    st.markdown("### 🔄 Return Analysis")
    returns_df = df[df['return_count'] > 0]

    if returns_df.empty:
        st.success("✅ Return topilmadi!")
        return

    col1, col2 = st.columns(2)
    with col1:
        st.metric("Jami Returns", int(returns_df['return_count'].sum()))
    with col2:
        st.metric("Tasks with Returns", len(returns_df))


def render_components_tab(df):
    """Components tab"""
    st.markdown("### 📦 Component Analysis")
    st.info("Component tahlili - under development")


def render_timeline_tab(df):
    """Timeline tab"""
    st.markdown("### 📅 Timeline Analysis")
    st.info("Timeline tahlili - under development")


# ============================================================================
# MAIN FUNCTION - OPTIMIZED FLOW
# ============================================================================

def render_statistics():
    """Statistics sahifasi - OPTIMIZED VERSION"""

    debug_log("=== STATISTICS SAHIFASI BOSHLANDI ===")

    # Header
    st.markdown("""
    <div class="main-header">
        <h1>📊 Sprint Statistics Dashboard</h1>
        <p>Jira sprint ma'lumotlarining batafsil tahlili</p>
        <p class="author">Developed by JASUR TURGUNOV</p>
    </div>
    """, unsafe_allow_html=True)

    excel_dir = os.getenv('EXCEL_DIR', './data/excel_reports')
    debug_log(f"Excel papka: {excel_dir}")

    # STEP 1: Get sprint metadata (FAST!)
    with st.spinner("🔍 Sprintlar izlanmoqda..."):
        metadata = get_sprint_metadata(excel_dir)

    if not metadata:
        st.error("❌ Excel fayllar topilmadi!")
        st.info(f"📂 Papka: {excel_dir}")
        return

    # Show sprint info
    st.success(f"✅ {len(metadata)} ta sprint topildi")

    # Sprint info table
    with st.expander("📋 Sprint Ma'lumotlari", expanded=False):
        sprint_info = []
        for sprint_id, meta in metadata.items():
            sprint_info.append({
                'Sprint ID': sprint_id,
                'Fayl': meta['file'],
                'Issues': meta['total_rows'],
                'Hajm': meta['file_size']
            })
        st.dataframe(pd.DataFrame(sprint_info), use_container_width=True, hide_index=True)

    st.markdown("---")

    # STEP 2: Sprint selection
    st.markdown("### 🎯 Sprint Tanlash")

    sprint_ids = sorted(metadata.keys())

    col1, col2, col3 = st.columns([3, 1, 1])

    with col1:
        selected_sprints = st.multiselect(
            "Sprint(lar)ni tanlang",
            sprint_ids,
            default=[],  # BO'SH - foydalanuvchi tanlaydi!
            help="⚡ Faqat tanlangan sprintlar yuklanadi (tezroq!)"
        )

    with col2:
        if st.button("🔄 Barchasini tanlash", use_container_width=True):
            selected_sprints = sprint_ids
            st.rerun()

    with col3:
        if st.button("🗑️ Tozalash", use_container_width=True):
            selected_sprints = []
            st.rerun()

    if not selected_sprints:
        st.info("👆 Sprintlarni tanlang va tahlilni boshlang!")
        return

    # Show selected info
    total_issues = sum(metadata[sid]['total_rows'] for sid in selected_sprints)
    st.info(f"📊 Tanlandi: {len(selected_sprints)} ta sprint, ~{total_issues} ta issue")

    st.markdown("---")

    # STEP 3: Load selected sprints (LAZY!)
    with st.spinner(f"⏳ {len(selected_sprints)} ta sprint yuklanmoqda..."):
        df = load_selected_sprints(metadata, selected_sprints)

    if df.empty:
        st.error("❌ Ma'lumot yuklanmadi!")
        return

    st.success(f"✅ Yuklandi: {len(df)} ta issue")

    st.markdown("---")

    # STEP 4: Statistics
    render_overview_metrics(df)

    st.markdown("---")

    # STEP 5: Tabs
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "👥 Developers",
        "🐛 Bugs",
        "📦 Components",
        "🔄 Returns",
        "📅 Timeline"
    ])

    with tab1:
        render_developers_tab(df)
    with tab2:
        render_bugs_tab(df)
    with tab3:
        render_components_tab(df)
    with tab4:
        render_returns_tab(df)
    with tab5:
        render_timeline_tab(df)

    debug_log("=== STATISTICS SAHIFASI TUGADI ===")